"""
分析結果をExcelに出力する
"""

from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(profitable_items: list, output_path: str = None) -> str:
    if output_path is None:
        output_path = f"kankocho_profit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "利益商品リスト"

    # ヘッダー
    headers = [
        "商品名",
        "出品価格（円）",
        "推定相場（円）",
        "推定利益（円）",
        "利益率",
        "相場ソース",
        "オークション種別",
        "主催機関",
        "商品説明",
        "備考（ブランド等）",
        "商品URL",
        "お気に入り登録",
    ]

    # ヘッダースタイル
    header_fill = PatternFill(start_color="F78C23", end_color="F78C23", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # データ行
    for row_idx, item in enumerate(profitable_items, 2):
        profit = item.get("estimated_profit", 0)
        auction_price = item.get("auction_price", 0)
        estimated_price = item.get("estimated_price", 0)
        profit_rate = (profit / auction_price * 100) if auction_price > 0 else 0

        sources_str = ", ".join(item.get("price_sources", []))

        note = ""
        if item.get("is_brand"):
            note = f"⚠️ ブランド品（{item.get('brand_name', '')}）：真贋判定未実施"

        values = [
            item.get("title", ""),
            auction_price,
            estimated_price,
            profit,
            f"{profit_rate:.1f}%",
            sources_str,
            item.get("auction_type", ""),
            item.get("organizer", ""),
            item.get("description", "")[:200],
            note,
            item.get("url", ""),
            "済" if item.get("favorited") else "未",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # 利益列の色分け
            if col == 4:
                if profit >= 10000:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif profit >= 3000:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        ws.row_dimensions[row_idx].height = 60

    # 列幅の調整
    col_widths = [40, 15, 15, 15, 10, 25, 20, 25, 50, 35, 50, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 件数サマリー
    summary_row = len(profitable_items) + 3
    ws.cell(row=summary_row, column=1, value=f"合計 {len(profitable_items)} 件の利益商品")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    if profitable_items:
        total_profit = sum(item.get("estimated_profit", 0) for item in profitable_items)
        ws.cell(row=summary_row + 1, column=1,
                value=f"推定利益合計（全商品落札・売却した場合）: {total_profit:,} 円")

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path
